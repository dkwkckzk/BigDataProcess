#!/usr/bin/python3

from openpyxl import load_workbook


# 엑셀 파일 불러오기
wb = load_workbook('student.xlsx')  # 실제 파일 경로와 이름 입력
ws = wb.active

data = [] 
for row in ws.iter_rows(values_only=True): # 각 셀의 값만을 포함하는 튜플 반환.
    data.append(row) # 성적,등급 =None인 튜플들을 data 리스트에 저장

num_students = len(data) - 1 # 헤드 부분을 제외한 학생 수만을 저장

totals = []
for i in range(1, len(data)): # 학생 수 만큼 돔
    midterm, final, hw, attendance = data[i][2:6] # i에 저장되어 있는 값 중 midterm 부분 부터 attendance부분까지 풀어헤쳐 각 자리에 나누어 각 변수에 대입
    total = midterm*0.3 + final*0.35 + hw*0.34 + attendance*1 # total 계산
    
    totals.append((total, i)) # i번째 에 총점 리스트에 추가 즉 해당하는 학생의 총점을 저장

sorted_totals=sorted(totals,key=lambda x:x[0],reverse=True)
# totals 정렬할건데 key값은 첫번째 자리에 있는 총점을 기준으로 하고, 거꾸로 즉 내림차순으로 정렬해줘

#grades_limits = [int(num_students*x)-1 for x in [0.15, 0.30, 0.55, 0.70, 0.85, 1]]
#grades_limits = [round(num_students*x)-1 for x in [0.15, 0.30, 0.50, 0.70, 0.85, 1]]
grades_limits = [round(num_students*x) for x in [0.15, 0.30, 0.50, 0.70, 0.85, 1]]
# 각 비율에 해당하는 학생수 저장이기는 한데 이게 맞나...?

grades = ['A+', 'A0', 'B+', 'B0', 'C+', 'C0']  # 등급 수정

# for rank, (total2, i) in enumerate(sorted_totals): 
#     if total2 < 40:
#         grade = 'F'
#     else:
#         grade = 'C0' 
#         for limit, g in zip(grades_limits, grades):
#             if rank <= limit:
#                 grade = g
#                 break

for rank, (total2, i) in enumerate(sorted_totals): 
    if total2 < 40:
        grade = 'F'  # 40점 미만일 때 'F' 등급 부여
    else:
        for limit, g in zip(grades_limits, grades):
            if rank < limit:
                grade = g
                break

   
    ws.cell(row=i+1, column=7).value = total2
    ws.cell(row=i+1, column=8).value = grade
   
wb.save('student.xlsx')
